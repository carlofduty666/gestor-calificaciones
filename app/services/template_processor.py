from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
from app.models.templates import ExcelTemplate, TemplateCell, TemplateRange, TemplateStyle
from app.models.academic import Student, Section, Subject, Period, TeacherAssignment
from app.models.grades import FinalGrade, StudentGrade
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from app import db
from io import BytesIO
import json
import re
import os
from datetime import datetime

class TemplateProcessor:
    """Procesador de plantillas Excel con inyección de datos"""
    
    def __init__(self, template):
        self.template = template
        self.workbook = None
        self.worksheet = None
        
    def load_template(self):
        """Cargar la plantilla Excel"""
        if not self.template.file_path or not os.path.exists(self.template.file_path):
            raise Exception(f"Archivo de plantilla no encontrado: {self.template.file_path}")
        
        self.workbook = load_workbook(self.template.file_path)
        self.worksheet = self.workbook.active
        
    def generate_section_report(self, section, period):
        """Generar reporte de sección usando la plantilla"""
        
        self.load_template()
        
        # Obtener estudiantes y datos
        students = Student.query.filter_by(
            section_id=section.id,
            is_active=True
        ).order_by(Student.last_name).all()
        
        # Obtener asignaturas
        subject_ids = db.session.query(TeacherAssignment.subject_id).filter_by(
            section_id=section.id,
            academic_year_id=period.academic_year_id
        ).distinct().all()
        
        subject_ids = [s[0] for s in subject_ids]
        subjects = Subject.query.filter(Subject.id.in_(subject_ids)).order_by(Subject.name).all()
        
        # Preparar contexto de datos
        context = {
            'section': section,
            'period': period,
            'students': students,
            'subjects': subjects,
            'current_date': datetime.now(),
            'total_students': len(students)
        }
        
        # Procesar celdas individuales
        self._process_individual_cells(context)
        
        # Procesar rangos iterativos
        self._process_ranges(context)
        
        # Guardar en memoria
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output


    def process_template(self, context):
            """
            Método principal para procesar una plantilla con datos
            
            Args:
                context: Diccionario con los datos para llenar la plantilla
                
            Returns:
                openpyxl.Workbook: El libro de Excel procesado
            """
            try:
                # Cargar o crear workbook
                if self.template.file_path and os.path.exists(self.template.file_path):
                    self.workbook = load_workbook(self.template.file_path)
                    self.worksheet = self.workbook.active
                else:
                    self.workbook = Workbook()
                    self.worksheet = self.workbook.active
                    self.worksheet.title = "Reporte"
                
                # Procesar diferentes tipos de contenido
                self._process_individual_cells(context)
                self._process_student_table(context)
                self._process_individual_cells(context)
                self._process_ranges(context)
                self._process_subject_headers(context)
                self._apply_template_styles()
                
                return self.workbook
                
            except Exception as e:
                print(f"Error procesando plantilla: {e}")
                import traceback
                traceback.print_exc()
                raise Exception(f"Error procesando plantilla: {str(e)}")
    
    def _process_student_table(self, context):
        """Procesar tabla de estudiantes si existe"""
        try:
            students = context.get('students', [])
            subjects = context.get('subjects', [])
            grades_data = context.get('grades_data', {})
            
            if not students:
                return
            
            # Buscar celdas marcadas como tabla de estudiantes
            table_cells = TemplateCell.query.filter_by(
                template_id=self.template.id,
                content_type='student_table'
            ).all()
            
            if not table_cells:
                return
            
            # Usar la primera celda como punto de inicio
            start_cell = table_cells[0]
            start_row = int(''.join(filter(str.isdigit, start_cell.cell_address)))
            start_col = self._column_letter_to_number(
                ''.join(filter(str.isalpha, start_cell.cell_address))
            )
            
            # Escribir headers de estudiantes
            self.worksheet.cell(row=start_row, column=start_col, value="ID")
            self.worksheet.cell(row=start_row, column=start_col + 1, value="Apellidos")
            self.worksheet.cell(row=start_row, column=start_col + 2, value="Nombres")
            
            # Headers de asignaturas
            for i, subject in enumerate(subjects):
                self.worksheet.cell(
                    row=start_row, 
                    column=start_col + 3 + i, 
                    value=subject.name
                )
            
            # Datos de estudiantes
            for i, student in enumerate(students):
                row = start_row + 1 + i
                
                self.worksheet.cell(row=row, column=start_col, value=student.student_id)
                self.worksheet.cell(row=row, column=start_col + 1, value=student.last_name)
                self.worksheet.cell(row=row, column=start_col + 2, value=student.first_name)
                
                # Calificaciones
                for j, subject in enumerate(subjects):
                    grade = grades_data.get(student.id, {}).get(subject.id)
                    self.worksheet.cell(
                        row=row, 
                        column=start_col + 3 + j, 
                        value=grade if grade is not None else ""
                    )
                    
        except Exception as e:
            print(f"Error procesando tabla de estudiantes: {e}")
    
    def _process_subject_headers(self, context):
        """Procesar headers de asignaturas dinámicamente"""
        try:
            subjects = context.get('subjects', [])
            
            # Buscar celdas marcadas como headers de asignaturas
            header_cells = TemplateCell.query.filter_by(
                template_id=self.template.id,
                content_type='subject_header'
            ).all()
            
            for i, subject in enumerate(subjects):
                if i < len(header_cells):
                    cell = header_cells[i]
                    self.worksheet[cell.cell_address] = subject.name
                    
        except Exception as e:
            print(f"Error procesando headers de asignaturas: {e}")
    
    def _apply_template_styles(self):
        """Aplicar estilos globales de la plantilla"""
        try:
            # Aplicar estilos de rangos
            template_styles = TemplateStyle.query.filter_by(
                template_id=self.template.id
            ).all()
            
            for style in template_styles:
                self._apply_range_style(style)
                
        except Exception as e:
            print(f"Error aplicando estilos de plantilla: {e}")
    
    def _apply_range_style(self, template_style):
        """Aplicar estilo a un rango de celdas"""
        try:
            # Ancho de columna
            if template_style.column_width and ':' in template_style.range_address:
                col_range = template_style.range_address.split(':')
                for col in col_range:
                    if col.isalpha():
                        self.worksheet.column_dimensions[col].width = template_style.column_width
            
            # Alto de fila
            if template_style.row_height and template_style.range_address.isdigit():
                self.worksheet.row_dimensions[int(template_style.range_address)].height = template_style.row_height
                
        except Exception as e:
            print(f"Error aplicando estilo de rango: {e}")
    
    def _column_letter_to_number(self, column_letter):
        """Convierte letra de columna a número"""
        result = 0
        for char in column_letter:
            result = result * 26 + (ord(char.upper()) - ord('A') + 1)
        return result

    def _apply_cell_style(self, cell, style_config_json):
        """Aplicar estilo a una celda"""
        
        try:
            # DESERIALIZAR EL JSON PRIMERO
            if isinstance(style_config_json, str):
                style_config = json.loads(style_config_json)
            else:
                style_config = style_config_json
            
            # Fuente
            if 'font' in style_config:
                font_config = style_config['font']
                font_color = font_config.get('color', '000000')
                
                # Asegurar formato correcto de color
                if len(font_color) == 6 and all(c in '0123456789ABCDEFabcdef' for c in font_color):
                    cell.font = Font(
                        name=font_config.get('name', 'Arial'),
                        size=font_config.get('size', 11),
                        bold=font_config.get('bold', False),
                        italic=font_config.get('italic', False),
                        color=font_color
                    )
            
            # Relleno - Solo aplicar si existe y no es blanco
            if 'fill' in style_config:
                fill_config = style_config['fill']
                fill_color = fill_config.get('color', 'FFFFFF')
                
                # Validar color y aplicar solo si no es blanco
                if (len(fill_color) == 6 and 
                    all(c in '0123456789ABCDEFabcdef' for c in fill_color) and 
                    fill_color.upper() != 'FFFFFF'):
                    
                    cell.fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type='solid'
                    )
            
            # Alineación
            if 'alignment' in style_config:
                align_config = style_config['alignment']
                cell.alignment = Alignment(
                    horizontal=align_config.get('horizontal', 'left'),
                    vertical=align_config.get('vertical', 'top'),
                    wrap_text=align_config.get('wrap_text', False)
                )
            
            # Bordes - CORREGIDO
            if 'border' in style_config:
                border_config = style_config['border']
                
                # Extraer solo el estilo de cada lado (no todo el diccionario)
                left_side = None
                if 'left' in border_config and isinstance(border_config['left'], dict):
                    left_style = border_config['left'].get('style')
                    if left_style:
                        left_side = Side(style=left_style)
                
                right_side = None
                if 'right' in border_config and isinstance(border_config['right'], dict):
                    right_style = border_config['right'].get('style')
                    if right_style:
                        right_side = Side(style=right_style)
                
                top_side = None
                if 'top' in border_config and isinstance(border_config['top'], dict):
                    top_style = border_config['top'].get('style')
                    if top_style:
                        top_side = Side(style=top_style)
                
                bottom_side = None
                if 'bottom' in border_config and isinstance(border_config['bottom'], dict):
                    bottom_style = border_config['bottom'].get('style')
                    if bottom_style:
                        bottom_side = Side(style=bottom_style)
                
                # Solo aplicar si hay al menos un borde
                if any([left_side, right_side, top_side, bottom_side]):
                    cell.border = Border(
                        left=left_side,
                        right=right_side,
                        top=top_side,
                        bottom=bottom_side
                    )
                    
        except Exception as e:
            print(f"Error aplicando estilo a celda {cell.coordinate}: {e}")
            import traceback
            traceback.print_exc()
            # No aplicar ningún estilo si hay error
            pass


    def generate_preview(self):
        """Generar vista previa con datos de ejemplo"""
        
        # Crear contexto de ejemplo
        example_context = {
            'section': type('Section', (), {
                'grade': type('Grade', (), {'name': '1er Año'})(),
                'name': 'A'
            })(),
            'period': type('Period', (), {'name': '1er Lapso'})(),
            'current_date': datetime.now(),
            'total_students': 25,
            'student': type('Student', (), {
                'first_name': 'Juan',
                'last_name': 'Pérez',
                'student_id': '12345678'
            })()
        }
        
        self.load_template()
        
        # Procesar solo celdas individuales para la vista previa
        preview_data = {}
        
        individual_cells = TemplateCell.query.filter_by(
            template_id=self.template.id
        ).all()
        
        for cell in individual_cells:
            try:
                value = self._get_cell_value(cell, example_context)
                
                # Parsear dirección de celda
                match = re.match(r'([A-Z]+)(\d+)', cell.cell_address)
                if match:
                    col_letter = match.group(1)
                    row_num = int(match.group(2))
                    
                    if row_num not in preview_data:
                        preview_data[row_num] = {}
                    
                    preview_data[row_num][col_letter] = {
                        'value': value,
                        'type': cell.data_type
                    }
                    
            except Exception as e:
                continue
        
        return preview_data

    
    def generate_student_report(self, student, period):
        """Generar reporte individual de estudiante"""
        
        self.load_template()
        
        # Obtener asignaturas del estudiante
        subject_ids = db.session.query(TeacherAssignment.subject_id).filter_by(
            section_id=student.section_id,
            academic_year_id=period.academic_year_id
        ).distinct().all()
        
        subject_ids = [s[0] for s in subject_ids]
        subjects = Subject.query.filter(Subject.id.in_(subject_ids)).order_by(Subject.name).all()
        
        # Obtener calificaciones
        grades_data = {}
        for subject in subjects:
            final_grade = FinalGrade.query.filter_by(
                student_id=student.id,
                subject_id=subject.id,
                period_id=period.id
            ).first()
            
            if final_grade:
                grades_data[subject.id] = final_grade.value
        
        # Preparar contexto
        context = {
            'student': student,
            'period': period,
            'subjects': subjects,
            'grades_data': grades_data,
            'current_date': datetime.now()
        }
        
        # Procesar plantilla
        self._process_individual_cells(context)
        self._process_ranges(context)
        
        # Guardar en memoria
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output
    
    def _process_individual_cells(self, context):
        """Procesar celdas individuales"""
        
        individual_cells = TemplateCell.query.filter_by(
            template_id=self.template.id
        ).all()
        
        # print(f"=== DEBUG PROCESANDO CELDAS ===")
        # print(f"Template ID: {self.template.id}")
        # print(f"Total celdas encontradas: {len(individual_cells)}")
        
        for cell in individual_cells:
            # print(f"\nProcesando celda {cell.cell_address}:")
            # print(f"  - Data Type: '{cell.data_type}'")
            # print(f"  - Default Value: '{cell.default_value}'")
            # print(f"  - Cell Type: '{cell.cell_type}'")
            
            try:
                value = self._get_cell_value(cell, context)
                # print(f"  - Valor calculado: '{value}'")
                
                if value is not None and value != '':
                    self.worksheet[cell.cell_address] = value
                    # print(f"  - ✓ Valor '{value}' asignado a celda {cell.cell_address}")
                    
                    # Aplicar estilo si existe
                    if cell.style_config:
                        # print(f"  - Aplicando estilo...")
                        self._apply_cell_style(
                            self.worksheet[cell.cell_address], 
                            cell.style_config
                        )
                        # print(f"  - ✓ Estilo aplicado")
                    
            except Exception as e:
                print(f"  - ✗ Error procesando celda: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        print(f"=== FIN DEBUG CELDAS ===")
    
    def _process_ranges(self, context):
        """Procesar rangos iterativos"""
        
        ranges = TemplateRange.query.filter_by(
            template_id=self.template.id
        ).all()
        
        print(f"=== DEBUG _process_ranges ===")
        print(f"Template ID: {self.template.id}")
        print(f"Rangos encontrados: {len(ranges)}")
        
        for range_obj in ranges:
            print(f"Procesando rango: {range_obj.range_name}")
            print(f"  - Tipo: {range_obj.range_type}")
            print(f"  - Start: {range_obj.start_cell}")
            print(f"  - Mapping: {range_obj.data_mapping}")
            
            try:
                self._process_single_range(range_obj, context)
                print(f"  - ✓ Rango procesado exitosamente")
            except Exception as e:
                print(f"  - ✗ Error procesando rango: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        print(f"=== FIN DEBUG _process_ranges ===")

    def _process_students_range(self, range_obj, context):
        """Procesar rango de estudiantes"""
        
        import json
        import re
        from openpyxl.utils import column_index_from_string, get_column_letter
        
        # Obtener estudiantes del contexto
        students = context.get('students', [])
        print(f"    Estudiantes disponibles: {len(students)}")
        
        if not students:
            print("    No hay estudiantes para procesar")
            return
        
        # Parsear mapping
        try:
            mapping = json.loads(range_obj.data_mapping)
            print(f"    Mapping: {mapping}")
        except Exception as e:
            print(f"    Error parseando mapping: {e}")
            return
        
        # Procesar rango
        start_cell = range_obj.start_cell
        end_cell = range_obj.end_cell
        
        if not start_cell:
            print("    No hay celda inicial")
            return
        
        # Extraer coordenadas del rango
        start_match = re.match(r'([A-Z]+)(\d+)', start_cell.upper())  # ← Agregar .upper()
        if not start_match:
            print(f"    Formato de celda inicial inválido: {start_cell}")
            return
        
        start_col = start_match.group(1)
        start_row = int(start_match.group(2))
        
        # Si hay celda final, calcular rango completo
        if end_cell:
            end_match = re.match(r'([A-Z]+)(\d+)', end_cell.upper())  # ← Agregar .upper()
            if end_match:
                end_col = end_match.group(1)
                end_row = int(end_match.group(2))
                print(f"    Rango completo: {start_cell} hasta {end_cell}")
            else:
                end_col = start_col
                end_row = start_row + len(students) - 1
                print(f"    Rango calculado: {start_cell} hasta {end_col}{end_row}")
        else:
            # Rango abierto: solo columna inicial, tantas filas como estudiantes
            end_col = start_col
            end_row = start_row + len(students) - 1
            print(f"    Rango abierto: {start_cell} hasta {end_col}{end_row}")
        
        # Determinar columnas a procesar
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        print(f"    Columnas a procesar: {start_col} ({start_col_idx}) hasta {end_col} ({end_col_idx})")
        
        # Procesar cada estudiante
        for i, student in enumerate(students):
            current_row = start_row + i
            
            # Si excede el rango de filas, parar
            if current_row > end_row:
                print(f"    Límite de filas alcanzado en fila {current_row}")
                break
            
            # ← CAMBIAR ESTA LÍNEA:
            print(f"      Estudiante {i+1}: {student.first_name} {student.last_name} -> Fila {current_row}")
            
            # Procesar cada columna del rango
            for col_idx in range(start_col_idx, end_col_idx + 1):
                col_letter = get_column_letter(col_idx)
                cell_address = f"{col_letter}{current_row}"
                
                # Determinar qué dato va en esta columna
                data_type = self._get_data_type_for_column(col_letter, mapping, start_col)
                
                if data_type:
                    value = self._get_student_data_value(student, data_type, i)
                    print(f"        {cell_address} = '{value}' (tipo: {data_type})")
                    
                    if value is not None and value != '':
                        self.worksheet[cell_address] = value
                else:
                    print(f"        {cell_address} = (sin mapeo)")

    def _get_data_type_for_column(self, col_letter, mapping, start_col):
        """Determinar qué tipo de dato va en una columna específica"""
        
        # Nuevo sistema simplificado
        if 'tipo' in mapping:
            return mapping['tipo']
        
        # Fallback para sistema anterior
        if col_letter in mapping:
            return mapping[col_letter]
        
        # Si es rango de una sola columna, usar el primer valor
        return list(mapping.values())[0] if mapping else None

    def _get_student_data_value(self, student, data_type, index=0):
        """Obtener valor de dato del estudiante"""
        
        # ← CAMBIAR TODAS ESTAS LÍNEAS:
        if data_type == 'cedula':
            return student.student_id or ''
        elif data_type == 'nombre_completo':
            first_name = student.first_name or 'didnt work bro'
            last_name = student.last_name or 'didnt work bro'
            return f"{first_name} {last_name}".strip()
        elif data_type == 'nombres':
            return student.first_name or 'didnt work bro'
        elif data_type == 'apellidos':
            return student.last_name or 'didnt work bro'
        elif data_type == 'numero_correlativo' or data_type == 'numero':
            return index + 1
        else:
            print(f"        Tipo de dato no reconocido: {data_type}")
            return ''
    
    def _process_single_range(self, range_obj, context):
        """Procesar un rango específico"""
        
        print(f"  _process_single_range: {range_obj.range_name}")
        print(f"    Tipo: {range_obj.range_type}")
        print(f"    Start cell: {range_obj.start_cell}")
        print(f"    Data mapping: {range_obj.data_mapping}")
        
        if range_obj.range_type == 'students':
            self._process_students_range(range_obj, context)
    
    def _get_students_data(self, context, column_mapping):
        """Obtener datos de estudiantes"""
        
        students = context.get('students', [])
        data_rows = []
        
        for index, student in enumerate(students):
            row_data = {
                'numero': index + 1,
                'cedula': student.student_id,
                'apellido': student.last_name,
                'nombre': student.first_name,
                'nombre_completo': f"{student.last_name}, {student.first_name}",
                'seccion': f"{student.section.grade.name}{student.section.name}"
            }
            data_rows.append(row_data)
        
        return data_rows
    
    def _get_subjects_data(self, context, column_mapping):
        """Obtener datos de asignaturas"""
        
        subjects = context.get('subjects', [])
        data_rows = []
        
        for index, subject in enumerate(subjects):
            row_data = {
                'numero': index + 1,
                'codigo': subject.code,
                'nombre': subject.name,
                'descripcion': subject.description or '',
                'profesor': self._get_subject_teacher(subject, context)
            }
            data_rows.append(row_data)
        
        return data_rows
    
    def _get_grades_data(self, context, column_mapping):
        """Obtener datos de calificaciones por estudiante"""
        
        students = context.get('students', [])
        subjects = context.get('subjects', [])
        period = context.get('period')
        data_rows = []
        
        for index, student in enumerate(students):
            row_data = {
                'numero': index + 1,
                'cedula': student.student_id,
                'nombre': f"{student.last_name}, {student.first_name}",
                'apellido': student.last_name,
                'nombre_solo': student.first_name
            }
            
            # Agregar calificaciones por materia
            total_grades = 0
            grade_count = 0
            
            for subject_index, subject in enumerate(subjects):
                final_grade = FinalGrade.query.filter_by(
                    student_id=student.id,
                    subject_id=subject.id,
                    period_id=period.id
                ).first()
                
                grade_value = final_grade.value if final_grade else 0
                row_data[f'materia_{subject_index + 1}'] = grade_value
                row_data[f'nota_{subject.code.lower()}'] = grade_value
                
                if grade_value > 0:
                    total_grades += grade_value
                    grade_count += 1
            
            # Calcular promedio
            row_data['promedio'] = round(total_grades / grade_count, 2) if grade_count > 0 else 0
            row_data['total_materias'] = len(subjects)
            row_data['materias_aprobadas'] = sum(1 for s in subjects if self._get_student_grade(student, s, period) >= 10)
            
            data_rows.append(row_data)
        
        return data_rows
    
    def _get_subject_teacher(self, subject, context):
        """Obtener el profesor de una materia"""
        
        section = context.get('section')
        period = context.get('period')
        
        if not section or not period:
            return ''
        
        from app.models.academic import TeacherAssignment, Teacher
        
        assignment = TeacherAssignment.query.filter_by(
            subject_id=subject.id,
            section_id=section.id,
            academic_year_id=period.academic_year_id
        ).first()
        
        if assignment and assignment.teacher:
            teacher_user = assignment.teacher.user
            return f"{teacher_user.first_name} {teacher_user.last_name}"
        
        return ''
    
    def _get_student_grade(self, student, subject, period):
        """Obtener calificación de un estudiante en una materia"""
        
        final_grade = FinalGrade.query.filter_by(
            student_id=student.id,
            subject_id=subject.id,
            period_id=period.id
        ).first()
        
        return final_grade.value if final_grade else 0
    
    def _get_cell_value(self, cell, context):
        """Obtener el valor para una celda según su tipo"""
        
        data_type = cell.data_type
        # print(f"    _get_cell_value: celda={cell.cell_address}, data_type='{data_type}', default_value='{cell.default_value}'")
        
        if data_type == 'static':
            result = cell.default_value
            # print(f"    → Retornando static: '{result}'")
            return result
        
        elif data_type == 'custom_text':
            result = cell.default_value
            # print(f"    → Retornando custom_text: '{result}'")
            return result
        
        elif data_type == 'titulo_reporte':
            result = 'REPORTE DE CALIFICACIONES'
            # print(f"    → Retornando titulo_reporte: '{result}'")
            return result
        
        elif data_type == 'nombre_institucion':
            return 'UNIDAD EDUCATIVA'  # Puedes hacer esto configurable
        
        elif data_type == 'seccion':
            section = context.get('section')
            return f"{section.grade.name} \"{section.name}\"" if section else ''
        
        elif data_type == 'grado':
            section = context.get('section')
            return section.grade.name if section else ''
        
        elif data_type == 'periodo':
            period = context.get('period')
            return period.name if period else ''
        
        elif data_type == 'fecha_actual':
            current_date = context.get('current_date')
            # Manejar tanto string como datetime
            if isinstance(current_date, str):
                return current_date
            elif hasattr(current_date, 'strftime'):
                return current_date.strftime('%d/%m/%Y')
            else:
                return ''
        
        elif data_type == 'total_estudiantes':
            return context.get('total_students', 0)
        
        elif data_type == 'promedio_seccion':
            return self._calculate_section_average(context)
        
        elif data_type == 'estudiantes_aprobados':
            return self._count_passing_students(context)
        
        elif data_type == 'estudiantes_reprobados':
            total = context.get('total_students', 0)
            passed = self._count_passing_students(context)
            return total - passed
        
        # Para estudiante individual
        elif data_type == 'estudiante_nombre':
            student = context.get('student')
            return f"{student.first_name} {student.last_name}" if student else ''
        
        elif data_type == 'estudiante_cedula':
            student = context.get('student')
            return student.student_id if student else ''
        
        elif data_type == 'estudiante_promedio':
            return self._calculate_student_average(context)
        
        # Si no coincide con ningún tipo específico, usar default_value
        result = cell.default_value
        # print(f"    → Retornando default: '{result}'")
        return result
    
    def _calculate_section_average(self, context):
        """Calcular promedio general de la sección"""
        
        students = context.get('students', [])
        subjects = context.get('subjects', [])
        period = context.get('period')
        
        if not students or not subjects or not period:
            return 0
        
        total_grades = 0
        grade_count = 0
        
        for student in students:
            for subject in subjects:
                grade = self._get_student_grade(student, subject, period)
                if grade > 0:
                    total_grades += grade
                    grade_count += 1
        
        return round(total_grades / grade_count, 2) if grade_count > 0 else 0
    
    def _count_passing_students(self, context):
        """Contar estudiantes aprobados"""
        
        students = context.get('students', [])
        subjects = context.get('subjects', [])
        period = context.get('period')
        
        if not students or not subjects or not period:
            return 0
        
        passing_students = 0
        
        for student in students:
            student_grades = []
            for subject in subjects:
                grade = self._get_student_grade(student, subject, period)
                if grade > 0:
                    student_grades.append(grade)
            
            if student_grades:
                average = sum(student_grades) / len(student_grades)
                if average >= 10:  # Nota mínima aprobatoria
                    passing_students += 1
        
        return passing_students
    
    def _calculate_student_average(self, context):
        """Calcular promedio de un estudiante"""
        
        student = context.get('student')
        grades_data = context.get('grades_data', {})
        
        if not student or not grades_data:
            return 0
        
        grades = [grade for grade in grades_data.values() if grade > 0]
        return round(sum(grades) / len(grades), 2) if grades else 0
    
    def _parse_cell_address(self, cell_address):
        """Parsear dirección de celda (ej: A1 -> (1, 1))"""
        
        match = re.match(r'([A-Z]+)(\d+)', cell_address)
        if not match:
            return 1, 1
        
        col_letters = match.group(1)
        row_num = int(match.group(2))
        
        # Convertir letras de columna a número
        col_num = 0
        for char in col_letters:
            col_num = col_num * 26 + (ord(char) - ord('A') + 1)
        
        return row_num, col_num
    

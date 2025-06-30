from openpyxl import load_workbook
from app.models.templates import ExcelTemplate, TemplateCell, TemplateStyle, TemplateRange
from app.services.template_mapper import TemplateMapper
from app import db
import json
import os
import re

class TemplateService:
    
    @staticmethod
    def analyze_template_structure(template_id, file_path):
        """Analiza la estructura de un archivo Excel y crea celdas automáticamente"""
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Archivo no encontrado: {file_path}")
        
        try:
            # Cargar workbook
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Limpiar celdas existentes
            TemplateCell.query.filter_by(template_id=template_id).delete()
            TemplateStyle.query.filter_by(template_id=template_id).delete()
            
            # print(f"🔍 Analizando estructura del template {template_id}")
            
            # Analizar celdas con contenido
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # 🔢 DETERMINAR TIPO DE DATO Y FORMATO
                        cell_data_type = TemplateService._determine_cell_data_type(cell)
                        formatted_value = TemplateService._format_cell_value(cell)
                        
                        # print(f"Celda: {cell.coordinate}, Valor: {cell.value}, Tipo: {cell_data_type}, Formateado: {formatted_value}")
                        
                        # Crear celda de plantilla
                        template_cell = TemplateCell(
                            template_id=template_id,
                            cell_address=cell.coordinate,
                            cell_type='static',  # Por defecto estático
                            default_value=formatted_value,  # Usar valor formateado
                            data_type=cell_data_type,  # 🔢 NUEVO: Guardar tipo de dato
                            style_config=TemplateService._extract_cell_style(cell)
                        )
                        db.session.add(template_cell)
            
            # Analizar estilos de columnas
            for col_letter, col_dimension in ws.column_dimensions.items():
                if col_dimension.width:
                    try:
                        template_style = TemplateStyle(
                            template_id=template_id,
                            range_address=f"{col_letter}:{col_letter}",
                            column_width=col_dimension.width
                        )
                        db.session.add(template_style)
                    except Exception as e:
                        print(f"Error procesando columna {col_letter}: {e}")
                        continue
            
            # Analizar estilos de filas
            for row_num, row_dimension in ws.row_dimensions.items():
                if row_dimension.height:
                    try:
                        template_style = TemplateStyle(
                            template_id=template_id,
                            range_address=f"{row_num}:{row_num}",
                            row_height=row_dimension.height
                        )
                        db.session.add(template_style)
                    except Exception as e:
                        print(f"Error procesando fila {row_num}: {e}")
                        continue
            
            db.session.commit()
            print(f"Plantilla {template_id} analizada exitosamente")
            
        except Exception as e:
            db.session.rollback()
            raise Exception(f"Error al analizar plantilla: {str(e)}")
        
    @staticmethod
    def _determine_cell_data_type(cell):
        """Determina el tipo de dato de una celda"""
        
        if cell.value is None:
            return 'text'
        
        # Verificar si es número
        if isinstance(cell.value, (int, float)):
            # Verificar formato de número en Excel
            if cell.number_format:
                format_code = cell.number_format.lower()
                
                # Formatos que indican decimales
                decimal_formats = ['0.0', '0.00', '#.#', '#.##', 'general', '0.000']
                
                # Formatos que indican enteros
                integer_formats = ['0', '#', '0_', '#_']
                
                # Si tiene formato específico de decimal
                if any(fmt in format_code for fmt in decimal_formats) and '.' in format_code:
                    return 'decimal'
                
                # Si tiene formato específico de entero
                if any(fmt in format_code for fmt in integer_formats):
                    return 'integer'
                
                # Si es porcentaje
                if '%' in format_code:
                    return 'percentage'
            
            # Si no hay formato específico, determinar por el valor
            if isinstance(cell.value, int):
                return 'integer'
            elif isinstance(cell.value, float):
                # Si el float es realmente un entero (ej: 15.0)
                if cell.value.is_integer():
                    return 'integer'
                else:
                    return 'decimal'
        
        # Verificar si es fecha
        if hasattr(cell, 'is_date') and cell.is_date:
            return 'date'
        
        # Por defecto es texto
        return 'text'

    @staticmethod
    def _format_cell_value(cell):
        """Formatea el valor de la celda preservando su tipo original"""
        
        if cell.value is None:
            return ''
        
        # Determinar tipo
        data_type = TemplateService._determine_cell_data_type(cell)
        
        if data_type == 'integer':
            # Preservar como entero
            if isinstance(cell.value, float) and cell.value.is_integer():
                return str(int(cell.value))
            elif isinstance(cell.value, int):
                return str(cell.value)
        
        elif data_type == 'decimal':
            # Preservar decimales
            if isinstance(cell.value, (int, float)):
                # Determinar número de decimales del formato original
                if cell.number_format and '.' in cell.number_format:
                    decimal_places = cell.number_format.count('0') - cell.number_format.find('.') - 1
                    if decimal_places > 0:
                        return f"{cell.value:.{decimal_places}f}"
                return str(float(cell.value))
        
        elif data_type == 'percentage':
            # Manejar porcentajes
            return f"{cell.value * 100}%" if isinstance(cell.value, float) else str(cell.value)
        
        # Para texto y otros tipos
        return str(cell.value)

    @staticmethod
    def _extract_cell_style(cell):
        """Extrae los estilos de una celda"""
        style_config = {}
        
        # print(f"🔍 DEBUG: Extrayendo estilo de celda {cell.coordinate}")
        
        try:
            # Fuente
            if cell.font:
                font_color = '000000'  # Default negro
                
                # ARREGLAR EXTRACCIÓN DE COLOR DE FUENTE
                try:
                    if cell.font.color and hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                        color_value = str(cell.font.color.rgb)
                        # print(f"   🎨 Color raw de fuente: {color_value}")
                        
                        # Manejar diferentes formatos
                        if color_value.startswith('FF') and len(color_value) == 8:
                            font_color = color_value[2:]  # Quitar FF (alpha)
                        elif len(color_value) == 6:
                            font_color = color_value
                        elif color_value == '0':
                            font_color = '000000'  # Negro
                        else:
                            font_color = '000000'  # Default si no se puede parsear
                            
                        # print(f"   📝 Color procesado de fuente: {font_color}")
                            
                except Exception as font_error:
                    print(f"   ❌ Error procesando color de fuente: {font_error}")
                    font_color = '000000'
                
                style_config['font'] = {
                    'name': cell.font.name or 'Arial',
                    'size': cell.font.size or 11,
                    'bold': cell.font.bold or False,
                    'italic': cell.font.italic or False,
                    'color': font_color
                }
                # print(f"   📝 Font extraída: {style_config['font']}")
            
            # ARREGLAR EXTRACCIÓN DE RELLENO
            if cell.fill and cell.fill.start_color:
                try:
                    fill_color = 'FFFFFF'  # Default blanco
                    
                    if hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
                        fill_value = str(cell.fill.start_color.rgb)
                        # print(f"   🎨 Fill raw: {fill_value}")
                        
                        # Manejar diferentes formatos
                        if fill_value.startswith('FF') and len(fill_value) == 8:
                            fill_color = fill_value[2:]  # Quitar FF (alpha)
                        elif len(fill_value) == 6:
                            fill_color = fill_value
                        elif fill_value == '0':
                            fill_color = 'FFFFFF'  # Blanco por defecto si es 0
                        else:
                            fill_color = 'FFFFFF'  # Default blanco
                            
                        # print(f"   🎨 Fill procesado: {fill_color}")
                    
                    # SOLO guardar si NO es blanco
                    if fill_color.upper() not in ['FFFFFF', 'FFFFFFFF']:
                        style_config['fill'] = {
                            'color': fill_color
                        }
                        # print(f"   ✅ Fill guardado: {fill_color}")
                    else:
                        print(f"   ⚪ Fill blanco ignorado")
                        
                except Exception as fill_error:
                    print(f"   ❌ Error procesando fill: {fill_error}")
            
            # BORDES - CAPTURAR EL ESTILO ORIGINAL EXACTO
            if cell.border:
                border_config = {}
                
                try:
                    # Capturar cada lado del borde con su estilo original
                    if cell.border.left and cell.border.left.style:
                        left_color = '000000'
                        try:
                            if cell.border.left.color and cell.border.left.color.rgb:
                                left_color = str(cell.border.left.color.rgb)
                                if left_color.startswith('FF') and len(left_color) == 8:
                                    left_color = left_color[2:]
                        except:
                            left_color = '000000'
                            
                        border_config['left'] = {
                            'style': cell.border.left.style,
                            'color': left_color
                        }
                    
                    if cell.border.right and cell.border.right.style:
                        right_color = '000000'
                        try:
                            if cell.border.right.color and cell.border.right.color.rgb:
                                right_color = str(cell.border.right.color.rgb)
                                if right_color.startswith('FF') and len(right_color) == 8:
                                    right_color = right_color[2:]
                        except:
                            right_color = '000000'
                            
                        border_config['right'] = {
                            'style': cell.border.right.style,
                            'color': right_color
                        }
                    
                    if cell.border.top and cell.border.top.style:
                        top_color = '000000'
                        try:
                            if cell.border.top.color and cell.border.top.color.rgb:
                                top_color = str(cell.border.top.color.rgb)
                                if top_color.startswith('FF') and len(top_color) == 8:
                                    top_color = top_color[2:]
                        except:
                            top_color = '000000'
                            
                        border_config['top'] = {
                            'style': cell.border.top.style,
                            'color': top_color
                        }
                    
                    if cell.border.bottom and cell.border.bottom.style:
                        bottom_color = '000000'
                        try:
                            if cell.border.bottom.color and cell.border.bottom.color.rgb:
                                bottom_color = str(cell.border.bottom.color.rgb)
                                if bottom_color.startswith('FF') and len(bottom_color) == 8:
                                    bottom_color = bottom_color[2:]
                        except:
                            bottom_color = '000000'
                            
                        border_config['bottom'] = {
                            'style': cell.border.bottom.style,
                            'color': bottom_color
                        }
                    
                    if border_config:
                        style_config['border'] = border_config
                        # print(f"   🔲 Border extraído: {border_config}")
                        
                except Exception as border_error:
                    print(f"   ❌ Error procesando bordes: {border_error}")
            
            # Alineación
            if cell.alignment:
                alignment_config = {
                    'horizontal': cell.alignment.horizontal or 'left',
                    'vertical': cell.alignment.vertical or 'top',
                    'wrap_text': cell.alignment.wrap_text or False
                }
                
                # 🔄 CAPTURAR ORIENTACIÓN DE TEXTO (VERTICAL/HORIZONTAL)
                if cell.alignment.text_rotation is not None:
                    alignment_config['text_rotation'] = cell.alignment.text_rotation
                    # print(f"   🔄 Rotación de texto detectada: {cell.alignment.text_rotation}°")
                
                # Capturar shrink_to_fit si existe
                if cell.alignment.shrink_to_fit is not None:
                    alignment_config['shrink_to_fit'] = cell.alignment.shrink_to_fit
                
                # Capturar indent si existe
                if cell.alignment.indent is not None and cell.alignment.indent > 0:
                    alignment_config['indent'] = cell.alignment.indent
                
                style_config['alignment'] = alignment_config
                # print(f"   📐 Alignment extraído: {alignment_config}")
            
            result = json.dumps(style_config) if style_config else None
            # print(f"   ✅ Estilo final para {cell.coordinate}: {result}")
            
            return result
            
        except Exception as e:
            print(f"❌ ERROR GENERAL extrayendo estilo de {cell.coordinate}: {e}")
            import traceback
            traceback.print_exc()
            return None

    # Agregar estos métodos al final de la clase TemplateService
    @staticmethod
    def generate_preview_with_config(template_id, config_data, sample_data):
        """Genera vista previa basada en configuración"""
        
        try:
            template = ExcelTemplate.query.get_or_404(template_id)
            preview_data = {}
            
            # Procesar celdas individuales
            if 'individual_cells' in config_data:
                for cell_config in config_data['individual_cells']:
                    address = cell_config['address']
                    data_type = cell_config['data_type']
                    custom_text = cell_config.get('custom_text', '')
                    
                    # Extraer fila y columna
                    row_num = int(''.join(filter(str.isdigit, address)))
                    col_letter = ''.join(filter(str.isalpha, address))
                    
                    if row_num not in preview_data:
                        preview_data[row_num] = {}
                    
                    # Generar valor según el tipo
                    value = TemplateService._get_preview_value(data_type, sample_data, custom_text)
                    
                    preview_data[row_num][col_letter] = {
                        'value': value,
                        'type': data_type
                    }

            # Procesar rangos iterativos
            if 'ranges' in config_data:
                for range_config in config_data['ranges']:
                    start_cell = range_config['start_cell']
                    mapping = range_config['mapping']
                    range_type = range_config['type']
                    
                    # Extraer posición inicial
                    start_row = int(''.join(filter(str.isdigit, start_cell)))
                    start_col_letter = ''.join(filter(str.isalpha, start_cell))
                    start_col_num = ord(start_col_letter) - ord('A')
                    
                    # Obtener datos según el tipo de rango
                    data_list = TemplateService._get_range_data(range_type, sample_data)
                    
                    # Llenar el rango
                    for row_offset, item_data in enumerate(data_list[:10]):  # Máximo 10 filas en preview
                        current_row = start_row + row_offset
                        
                        if current_row not in preview_data:
                            preview_data[current_row] = {}
                        
                        for col_offset, (col_key, data_key) in enumerate(mapping.items()):
                            current_col_num = start_col_num + col_offset
                            current_col_letter = chr(ord('A') + current_col_num)
                            
                            value = TemplateService._get_item_value(item_data, data_key, row_offset)
                            
                            preview_data[current_row][current_col_letter] = {
                                'value': value,
                                'type': data_key
                            }
            
            return preview_data
            
        except Exception as e:
            print(f"Error en generate_preview_with_config: {str(e)}")
            raise e

    @staticmethod
    def _get_preview_value(data_type, sample_data, custom_text=''):
        """Obtiene valor de vista previa para un tipo de dato"""
        
        if data_type == 'custom_text' and custom_text:
            return custom_text
        
        preview_values = {
            'static': 'Texto estático',
            'custom_text': custom_text or 'Texto personalizado',
            'titulo_reporte': 'REPORTE DE CALIFICACIONES',
            'nombre_institucion': 'U.E. MI INSTITUCIÓN',
            'seccion': sample_data.get('section', {}).get('name', '1ro "A"'),
            'grado': sample_data.get('section', {}).get('grade', {}).get('name', '1er Año'),
            'periodo': sample_data.get('period', {}).get('name', '1er Lapso'),
            'fecha_actual': sample_data.get('current_date', '13/06/2025'),
            'total_estudiantes': f"{len(sample_data.get('students', []))} estudiantes",
            'promedio_seccion': '16.8',
            'estudiantes_aprobados': '22 aprobados',
            'estudiantes_reprobados': '3 reprobados'
        }
        
        return preview_values.get(data_type, f'[{data_type}]')

    @staticmethod
    def _get_range_data(range_type, sample_data):
        """Obtiene datos para un tipo de rango"""
        
        if range_type == 'students':
            return sample_data.get('students', [])
        elif range_type == 'subjects':
            return sample_data.get('subjects', [])
        elif range_type == 'grades_by_student':
            students = sample_data.get('students', [])
            # Agregar notas de ejemplo
            for i, student in enumerate(students):
                student['materia_1'] = 18.5 - i
                student['materia_2'] = 17.0 + i * 0.5
                student['materia_3'] = 16.8 + i * 0.3
                student['promedio'] = (student['materia_1'] + student['materia_2'] + student['materia_3']) / 3
            return students
        
        return []

    @staticmethod
    def _get_item_value(item_data, data_key, index):
        """Obtiene valor específico de un item"""
        
        if data_key == 'numero':
            return str(index + 1)
        elif data_key in item_data:
            return str(item_data[data_key])
        else:
            return f'[{data_key}]'

    @staticmethod
    def generate_preview_excel(template_id):
        """Genera Excel de vista previa"""
        
        from openpyxl import Workbook
        
        # Por ahora retornamos un workbook básico
        wb = Workbook()
        ws = wb.active
        ws.title = "Vista Previa"
        
        ws['A1'] = "Vista previa de plantilla"
        ws['A2'] = "Funcionalidad en desarrollo"
        
        return wb

    @staticmethod
    def detect_data_ranges_from_template(template_id):
        """Detecta rangos de datos iterativos analizando las celdas existentes"""
        
        try:
            # Obtener todas las celdas del template
            cells = TemplateCell.query.filter_by(template_id=template_id).all()
            
            # Agrupar por filas para detectar patrones
            rows_data = {}
            for cell in cells:
                row_num = int(''.join(filter(str.isdigit, cell.cell_address)))
                col_letter = ''.join(filter(str.isalpha, cell.cell_address))
                
                if row_num not in rows_data:
                    rows_data[row_num] = {}
                
                rows_data[row_num][col_letter] = {
                    'data_type': cell.data_type,
                    'cell_type': cell.cell_type,
                    'style_config': cell.style_config
                }
            
            # Detectar rangos que parecen ser listas de estudiantes
            student_ranges = TemplateService._detect_student_ranges(rows_data)
            
            # Crear/actualizar TemplateRange entries
            for range_info in student_ranges:
                existing_range = TemplateRange.query.filter_by(
                    template_id=template_id,
                    range_name=range_info['name']
                ).first()
                
                if existing_range:
                    # Actualizar existente
                    existing_range.start_cell = range_info['start_cell']
                    existing_range.end_cell = range_info['end_cell']
                    existing_range.data_mapping = json.dumps(range_info['mapping'])
                else:
                    # Crear nuevo
                    new_range = TemplateRange(
                        template_id=template_id,
                        range_name=range_info['name'],
                        start_cell=range_info['start_cell'],
                        end_cell=range_info['end_cell'],
                        range_type='students',
                        data_mapping=json.dumps(range_info['mapping'])
                    )
                    db.session.add(new_range)
            
            db.session.commit()
            # print(f"✅ Rangos detectados y guardados para template {template_id}")
            return student_ranges
            
        except Exception as e:
            db.session.rollback()
            print(f"❌ Error detectando rangos: {e}")
            return []

    @staticmethod
    def _detect_student_ranges(rows_data):
        """Detecta rangos que contienen datos de estudiantes"""
        
        ranges = []
        sorted_rows = sorted(rows_data.keys())
        
        # Buscar secuencias consecutivas de filas con estructura similar
        current_sequence = []
        current_structure = None
        
        for row_num in sorted_rows:
            row_structure = TemplateService._get_row_data_types(rows_data[row_num])
            
            if current_structure is None:
                current_structure = row_structure
                current_sequence = [row_num]
            elif TemplateService._structures_similar(current_structure, row_structure):
                current_sequence.append(row_num)
            else:
                # Fin de secuencia - evaluar si es un rango de datos
                if len(current_sequence) >= 2:  # Al menos 2 filas
                    range_info = TemplateService._create_range_info(
                        current_sequence, current_structure, rows_data
                    )
                    if range_info:
                        ranges.append(range_info)
                
                # Iniciar nueva secuencia
                current_structure = row_structure
                current_sequence = [row_num]
        
        # No olvidar la última secuencia
        if len(current_sequence) >= 2:
            range_info = TemplateService._create_range_info(
                current_sequence, current_structure, rows_data
            )
            if range_info:
                ranges.append(range_info)
        
        return ranges

    @staticmethod
    def _get_row_data_types(row_data):
        """Extrae los tipos de datos de una fila"""
        structure = {}
        for col_letter, cell_data in row_data.items():
            structure[col_letter] = cell_data.get('data_type', 'text')
        return structure

    @staticmethod
    def _structures_similar(struct1, struct2):
        """Verifica si dos estructuras son similares (mismo patrón de columnas)"""
        return set(struct1.keys()) == set(struct2.keys())

    @staticmethod
    def _create_range_info(row_sequence, structure, rows_data):
        """Crea información del rango detectado"""
        
        start_row = min(row_sequence)
        end_row = max(row_sequence)
        
        # Determinar columnas del rango
        columns = sorted(structure.keys())
        start_col = columns[0] if columns else 'A'
        end_col = columns[-1] if columns else 'A'
        
        # Crear mapeo de columnas
        mapping = {}
        sample_row_data = rows_data[end_row]  # Usar última fila como muestra
        
        for col_letter in columns:
            cell_data = sample_row_data.get(col_letter, {})
            data_type = cell_data.get('data_type', 'text')
            mapping[col_letter] = data_type
        
        return {
            'name': f'students_range_{start_row}_{end_row}',
            'start_cell': f'{start_col}{start_row}',
            'end_cell': f'{end_col}{end_row}',
            'mapping': mapping,
            'sample_row': end_row,  # Fila que usaremos como plantilla para replicar
            'structure': structure
        }

    def _save_row_patterns(template_id, patterns):
        """Guarda los patrones detectados usando design_config existente"""
        
        template = ExcelTemplate.query.get(template_id)
        if template:
            template.set_row_patterns(patterns)  # Usa el método que agregamos
            db.session.commit()
            print(f"✅ Patrones guardados para template {template_id}")


    @staticmethod
    def extend_template_ranges(template_id, data):
        """Extiende los rangos del template según los datos requeridos"""
        
        try:
            template = ExcelTemplate.query.get(template_id)
            if not template:
                return False
            
            # Obtener patrones usando el método del modelo
            patterns = template.get_row_patterns()
            
            if not patterns:
                print("⚠️ No hay patrones de fila guardados")
                return False
            
        except Exception as e:
            db.session.rollback()
            print(f"❌ Error extendiendo rangos: {e}")
            return False

    @staticmethod
    def _extend_student_range(template_range, required_students):
        """Extiende un rango específico de estudiantes"""
        
        # Extraer información del rango actual
        start_row = int(''.join(filter(str.isdigit, template_range.start_cell)))
        end_row = int(''.join(filter(str.isdigit, template_range.end_cell))) if template_range.end_cell else start_row
        
        current_capacity = end_row - start_row + 1
        
        if required_students <= current_capacity:
            print(f"✅ Rango {template_range.range_name} ya tiene capacidad suficiente")
            return True
        
        # Necesitamos más filas
        additional_rows = required_students - current_capacity
        new_end_row = end_row + additional_rows
        
        print(f"🔄 Extendiendo rango {template_range.range_name}: {additional_rows} filas adicionales")
        
        # Obtener mapeo de columnas
        mapping = json.loads(template_range.data_mapping) if template_range.data_mapping else {}
        
        # Replicar estilos de la última fila para las nuevas filas
        for new_row in range(end_row + 1, new_end_row + 1):
            success = TemplateService._replicate_row_styles_from_range(
                template_range.template_id,
                end_row,  # Fila fuente (última fila existente)
                new_row,  # Fila destino
                mapping
            )
            
            if not success:
                print(f"❌ Error replicando fila {new_row}")
                return False
        
        # Actualizar el rango
        end_col = ''.join(filter(str.isalpha, template_range.end_cell)) if template_range.end_cell else 'A'
        template_range.end_cell = f"{end_col}{new_end_row}"
        
        return True

    @staticmethod
    def _replicate_row_styles_from_range(template_id, source_row, target_row, column_mapping):
        """Replica estilos de una fila a otra dentro de un rango"""
        
        try:
            # Obtener celdas de la fila fuente
            source_cells = TemplateCell.query.filter_by(template_id=template_id).all()
            
            for source_cell in source_cells:
                cell_row = int(''.join(filter(str.isdigit, source_cell.cell_address)))
                if cell_row != source_row:
                    continue
                
                col_letter = ''.join(filter(str.isalpha, source_cell.cell_address))
                
                # Solo replicar columnas que están en el mapeo
                if col_letter not in column_mapping:
                    continue
                
                target_address = f"{col_letter}{target_row}"
                
                # Verificar si ya existe
                existing_cell = TemplateCell.query.filter_by(
                    template_id=template_id,
                    cell_address=target_address
                ).first()
                
                if existing_cell:
                    # Actualizar existente
                    existing_cell.style_config = source_cell.style_config
                    existing_cell.data_type = column_mapping[col_letter]
                    existing_cell.cell_type = 'data'  # Marcar como celda de datos
                else:
                    # Crear nueva
                    new_cell = TemplateCell(
                        template_id=template_id,
                        cell_address=target_address,
                        cell_type='data',
                        data_type=column_mapping[col_letter],
                        default_value='',
                        style_config=source_cell.style_config
                    )
                    db.session.add(new_cell)
            
            return True
            
        except Exception as e:
            print(f"❌ Error replicando estilos: {e}")
            return False
        
    

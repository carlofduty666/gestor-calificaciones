from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from app.models.academic import AcademicYear, Period, Grade, Section
from app.services.data_import_service import DataImportService
from app.services.excel_data_extractor import ExcelDataExtractor
from app import db
import os
import tempfile
from datetime import datetime

data_import = Blueprint('data_import', __name__)

# Configuración de archivos permitidos
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@data_import.before_request
def check_access():
    """Verificar que el usuario tenga permisos para importar datos"""
    if not current_user.is_authenticated:
        return redirect(url_for('auth.login'))
    
    # Solo admins y profesores pueden importar datos
    if not (current_user.is_admin() or current_user.role == 'teacher'):
        flash('No tienes permisos para acceder a esta sección', 'danger')
        return redirect(url_for('teacher.dashboard'))

@data_import.route('/')
@login_required
def index():
    """Página principal de importación de datos"""
    
    # Obtener año académico activo
    active_year = AcademicYear.query.filter_by(is_active=True).first()
    
    if not active_year:
        flash('No hay un año académico activo', 'warning')
        return render_template('data_import/index.html', 
                             active_year=None, 
                             periods=[], 
                             sections=[])
    
    # Obtener períodos del año activo
    periods = active_year.periods.order_by(Period.start_date).all()
    
    # Obtener secciones disponibles
    sections = Section.query.join(Grade).order_by(Grade.name, Section.name).all()
    
    # Obtener historial de importaciones
    import_history = DataImportService.get_import_history(limit=10)
    
    return render_template('data_import/index.html',
                         active_year=active_year,
                         periods=periods,
                         sections=sections,
                         import_history=import_history)

@data_import.route('/upload', methods=['GET', 'POST'])
@login_required
def upload():
    """Subir archivo Excel para importación"""
    
    if request.method == 'POST':
        # Verificar que se haya subido un archivo
        if 'file' not in request.files:
            flash('No se seleccionó ningún archivo', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        section_id = request.form.get('section_id')
        period_id = request.form.get('period_id')
        
        if file.filename == '':
            flash('No se seleccionó ningún archivo', 'danger')
            return redirect(request.url)
        
        if not section_id or not period_id:
            flash('Debe seleccionar una sección y un período', 'danger')
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            # Guardar archivo temporalmente
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{filename}"
            
            # Crear directorio si no existe
            upload_dir = os.path.join('uploads', 'imports')
            os.makedirs(upload_dir, exist_ok=True)
            
            file_path = os.path.join(upload_dir, filename)
            file.save(file_path)
            
            # Validar archivo
            validation = ExcelDataExtractor.validate_excel_format(file_path)
            
            if not validation['is_valid']:
                # Eliminar archivo inválido
                os.remove(file_path)
                flash(f"Archivo inválido: {', '.join(validation['errors'])}", 'danger')
                return redirect(request.url)
            
            # Redirigir a vista previa
            return redirect(url_for('data_import.preview', 
                                  filename=filename, 
                                  section_id=section_id, 
                                  period_id=period_id))
        else:
            flash('Tipo de archivo no permitido. Solo se permiten archivos .xlsx y .xls', 'danger')
    
    # GET request - mostrar formulario
    active_year = AcademicYear.query.filter_by(is_active=True).first()
    
    if not active_year:
        flash('No hay un año académico activo', 'warning')
        return redirect(url_for('data_import.index'))
    
    periods = active_year.periods.order_by(Period.start_date).all()
    sections = Section.query.join(Grade).order_by(Grade.name, Section.name).all()
    
    return render_template('data_import/upload.html',
                         periods=periods,
                         sections=sections)

@data_import.route('/preview/<filename>')
@login_required
def preview(filename):
    """Vista previa de los datos a importar"""
    
    section_id = request.args.get('section_id')
    period_id = request.args.get('period_id')
    
    if not section_id or not period_id:
        flash('Parámetros faltantes', 'danger')
        return redirect(url_for('data_import.upload'))
    
    # Verificar que el archivo existe
    file_path = os.path.join('uploads', 'imports', filename)
    if not os.path.exists(file_path):
        flash('Archivo no encontrado', 'danger')
        return redirect(url_for('data_import.upload'))
    
    # Obtener sección y período
    section = Section.query.get_or_404(section_id)
    period = Period.query.get_or_404(period_id)
    
    # Analizar estructura del archivo
    try:
        structure_analysis = ExcelDataExtractor.analyze_excel_structure(file_path)
        preview_data = DataImportService.preview_import_data(file_path)
        
        return render_template('data_import/preview.html',
                             filename=filename,
                             section=section,
                             period=period,
                             structure_analysis=structure_analysis,
                             preview_data=preview_data)
    
    except Exception as e:
        flash(f'Error al analizar el archivo: {str(e)}', 'danger')
        return redirect(url_for('data_import.upload'))

@data_import.route('/process/<filename>', methods=['POST'])
@login_required
def process_import(filename):
    """Procesar la importación de datos"""
    
    section_id = request.form.get('section_id')
    period_id = request.form.get('period_id')
    sheet_name = request.form.get('sheet_name')
    
    if not section_id or not period_id:
        flash('Parámetros faltantes', 'danger')
        return redirect(url_for('data_import.upload'))
    
    # Verificar que el archivo existe
    file_path = os.path.join('uploads', 'imports', filename)
    if not os.path.exists(file_path):
        flash('Archivo no encontrado', 'danger')
        return redirect(url_for('data_import.upload'))
    
    try:
        # Procesar importación
        result = DataImportService.import_students_and_grades(
            file_path=file_path,
            section_id=section_id,
            period_id=period_id
        )
        
        # Limpiar archivo temporal
        os.remove(file_path)
        
        return render_template('data_import/result.html', result=result)
    
    except Exception as e:
        flash(f'Error durante la importación: {str(e)}', 'danger')
        return redirect(url_for('data_import.upload'))

@data_import.route('/mapping/<filename>')
@login_required
def mapping(filename):
    """Configurar mapeo de columnas (funcionalidad avanzada)"""
    
    section_id = request.args.get('section_id')
    period_id = request.args.get('period_id')
    
    # Verificar archivo
    file_path = os.path.join('uploads', 'imports', filename)
    if not os.path.exists(file_path):
        flash('Archivo no encontrado', 'danger')
        return redirect(url_for('data_import.upload'))
    
    try:
        # Analizar estructura
        structure_analysis = ExcelDataExtractor.analyze_excel_structure(file_path)
        
        # Obtener sección y período
        section = Section.query.get_or_404(section_id)
        period = Period.query.get_or_404(period_id)
        
        return render_template('data_import/mapping.html',
                             filename=filename,
                             section=section,
                             period=period,
                             structure_analysis=structure_analysis)
    
    except Exception as e:
        flash(f'Error al analizar archivo: {str(e)}', 'danger')
        return redirect(url_for('data_import.upload'))

@data_import.route('/api/validate-file', methods=['POST'])
@login_required
def validate_file_api():
    """API para validar archivo Excel"""
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'})
    
    file = request.files['file']
    
    if not file or file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'})
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'Invalid file type'})
    
    try:
        # Guardar temporalmente para validación
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            file.save(temp_file.name)
            
            # Validar
            validation = ExcelDataExtractor.validate_excel_format(temp_file.name)
            
            # Limpiar archivo temporal
            os.unlink(temp_file.name)
            
            return jsonify({
                'success': validation['is_valid'],
                'validation': validation
            })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@data_import.route('/download-template')
@login_required
def download_template():
    """Descargar plantilla Excel de ejemplo"""
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        # Crear workbook de ejemplo
        wb = Workbook()
        ws = wb.active
        ws.title = "Estudiantes y Calificaciones"
        
        # Encabezados
        headers = ['ID', 'Apellidos', 'Nombres', 'Matemáticas', 'Español', 'Ciencias', 'Historia']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Datos de ejemplo
        sample_data = [
            ['001', 'García', 'Juan', 85, 90, 78, 88],
            ['002', 'López', 'María', 92, 88, 95, 90],
            ['003', 'Martínez', 'Carlos', 78, 85, 82, 86],
            ['004', 'Rodríguez', 'Ana', 95, 92, 89, 94],
            ['005', 'Hernández', 'Luis', 80, 83, 87, 85]
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en archivo temporal
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            wb.save(temp_file.name)
            
            return send_file(
                temp_file.name,
                as_attachment=True,
                download_name='plantilla_importacion.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    
    except Exception as e:
        flash(f'Error al generar plantilla: {str(e)}', 'danger')
        return redirect(url_for('data_import.index'))

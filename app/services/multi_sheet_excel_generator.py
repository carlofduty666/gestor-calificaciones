from app.services.excel_generator import ExcelGenerator
from openpyxl import Workbook
from io import BytesIO
import tempfile

class MultiSheetExcelGenerator:
    
    @staticmethod
    def generate_section_report_with_template(template_id, section, period, subjects, students, grades_data):
        """Genera reporte de sección usando una plantilla"""
        
        # Preparar datos para la plantilla
        template_data = {
            'section': {
                'name': section.name,
                'grade': {
                    'name': section.grade.name
                }
            },
            'period': {
                'name': period.name
            },
            'current_date': period.academic_year.name,
            'subjects': [{'name': subject.name} for subject in subjects],
            'students': [],
            'grades_data': grades_data
        }
        
        # Agregar datos de estudiantes
        for student in students:
            student_data = {
                'id': student.student_id,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'full_name': f"{student.first_name} {student.last_name}"
            }
            template_data['students'].append(student_data)
        
        # Generar Excel usando la plantilla
        wb = ExcelGenerator.generate_from_template(template_id, template_data)
        
        # Agregar hoja con datos de calificaciones
        MultiSheetExcelGenerator._add_grades_sheet(wb, students, subjects, grades_data)
        
        return wb
    
    @staticmethod
    def _add_grades_sheet(wb, students, subjects, grades_data):
        """Agrega una hoja con los datos de calificaciones"""
        
        # Crear nueva hoja
        ws = wb.create_sheet("Calificaciones")
        
        # Encabezados
        headers = ['ID', 'Apellidos', 'Nombres'] + [subject.name for subject in subjects]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos de estudiantes y calificaciones
        for row, student in enumerate(students, 2):
            ws.cell(row=row, column=1, value=student.student_id)
            ws.cell(row=row, column=2, value=student.last_name)
            ws.cell(row=row, column=3, value=student.first_name)
            
            # Calificaciones por materia
            for col, subject in enumerate(subjects, 4):
                grade = grades_data.get(student.id, {}).get(subject.id, '')
                ws.cell(row=row, column=col, value=grade)
        
        # Aplicar estilos básicos
        MultiSheetExcelGenerator._apply_basic_styles(ws, len(headers), len(students) + 1)
    
    @staticmethod
    def _apply_basic_styles(ws, num_cols, num_rows):
        """Aplica estilos básicos a una hoja"""
        
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        # Estilo para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Aplicar a la primera fila
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Bordes para toda la tabla
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, num_rows + 1):
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).border = thin_border
    
    @staticmethod
    def generate_to_file(wb):
        """Guarda el workbook en un archivo temporal y retorna la ruta"""
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            wb.save(temp_file.name)
            return temp_file.name
    
    @staticmethod
    def generate_to_bytes(wb):
        """Convierte el workbook a bytes para descarga directa"""
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def generate_student_report_with_template(template_id, student, period, subjects, grades_data, final_grades):
        """Genera reporte individual de estudiante usando plantilla"""
        
        # Preparar datos para la plantilla
        template_data = {
            'student': {
                'id': student.student_id,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'full_name': f"{student.first_name} {student.last_name}",
                'section': {
                    'name': student.section.name,
                    'grade': {
                        'name': student.section.grade.name
                    }
                }
            },
            'period': {
                'name': period.name
            },
            'current_date': period.academic_year.name,
            'subjects': [],
            'grades_data': grades_data,
            'final_grades': final_grades
        }
        
        # Agregar datos de materias con calificaciones
        for subject in subjects:
            subject_data = {
                'name': subject.name,
                'final_grade': final_grades.get(subject.id, ''),
                'detailed_grades': grades_data.get(subject.id, {})
            }
            template_data['subjects'].append(subject_data)
        
        # Calcular promedio
        final_grade_values = [grade for grade in final_grades.values() if grade]
        if final_grade_values:
            try:
                average = sum(float(grade) for grade in final_grade_values) / len(final_grade_values)
                template_data['average'] = round(average, 2)
            except (ValueError, TypeError):
                template_data['average'] = 'N/A'
        else:
            template_data['average'] = 'N/A'
        
        # Generar Excel usando la plantilla
        wb = ExcelGenerator.generate_from_template(template_id, template_data)
        
        # Agregar hoja detallada si hay calificaciones por tipo
        if any(grades_data.values()):
            MultiSheetExcelGenerator._add_detailed_grades_sheet(wb, student, subjects, grades_data, final_grades)
        
        return wb
    
    @staticmethod
    def _add_detailed_grades_sheet(wb, student, subjects, grades_data, final_grades):
        """Agrega hoja con calificaciones detalladas por tipo de evaluación"""
        
        ws = wb.create_sheet("Detalle de Calificaciones")
        
        # Título
        ws.cell(row=1, column=1, value=f"Detalle de Calificaciones - {student.first_name} {student.last_name}")
        ws.merge_cells('A1:E1')
        
        current_row = 3
        
        for subject in subjects:
            # Nombre de la materia
            ws.cell(row=current_row, column=1, value=subject.name)
            ws.merge_cells(f'A{current_row}:E{current_row}')
            current_row += 1
            
            # Encabezados de tipos de calificación
            ws.cell(row=current_row, column=1, value="Tipo de Evaluación")
            ws.cell(row=current_row, column=2, value="Calificación")
            ws.cell(row=current_row, column=3, value="Peso")
            current_row += 1
            
            # Calificaciones por tipo
            subject_grades = grades_data.get(subject.id, {})
            for grade_type_id, grade_info in subject_grades.items():
                ws.cell(row=current_row, column=1, value=grade_info.get('name', ''))
                ws.cell(row=current_row, column=2, value=grade_info.get('value', ''))
                ws.cell(row=current_row, column=3, value=f"{grade_info.get('weight', 0)}%")
                current_row += 1
            
            # Calificación final
            ws.cell(row=current_row, column=1, value="CALIFICACIÓN FINAL")
            ws.cell(row=current_row, column=2, value=final_grades.get(subject.id, ''))
            current_row += 2  # Espacio entre materias
        
        # Aplicar estilos
        MultiSheetExcelGenerator._apply_detailed_styles(ws, current_row)
    
    @staticmethod
    def _apply_detailed_styles(ws, max_row):
        """Aplica estilos a la hoja de calificaciones detalladas"""
        
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Estilo para título
        title_font = Font(size=14, bold=True)
        title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        title_alignment = Alignment(horizontal="center")
        
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = title_alignment
        
        # Estilos para encabezados de materias
        subject_font = Font(bold=True, color="FFFFFF")
        subject_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Estilos para encabezados de evaluaciones
        eval_font = Font(bold=True)
        eval_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        
        # Aplicar estilos dinámicamente (esto es básico, se puede mejorar)
        for row in range(1, max_row + 1):
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    if "Detalle de Calificaciones" in cell.value:
                        cell.font = title_font
                        cell.fill = title_fill
                        cell.alignment = title_alignment

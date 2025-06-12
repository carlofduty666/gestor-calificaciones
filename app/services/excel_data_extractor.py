from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
import re

class ExcelDataExtractor:
    """Servicio para extraer datos de archivos Excel y convertirlos en datos estructurados"""
    
    @staticmethod
    def extract_students_and_grades(file_path, sheet_name=None):
        """
        Extrae estudiantes y calificaciones de un archivo Excel
        Formato esperado: Primera fila = encabezados, Primera columna = ID, Segunda = Apellidos, Tercera = Nombres
        """
        
        try:
            # Cargar archivo Excel
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            # Validar estructura mínima
            if len(df.columns) < 3:
                raise ValueError("El archivo debe tener al menos 3 columnas: ID, Apellidos, Nombres")
            
            # Limpiar datos
            df = df.dropna(subset=[df.columns[0]])  # Eliminar filas sin ID
            
            extracted_data = {
                'students': [],
                'subjects': [],
                'grades': []
            }
            
            # Extraer información de columnas
            columns = df.columns.tolist()
            id_col = columns[0]
            lastname_col = columns[1] 
            firstname_col = columns[2]
            subject_cols = columns[3:]  # El resto son materias
            
            # Extraer materias
            for subject_col in subject_cols:
                if pd.notna(subject_col) and str(subject_col).strip():
                    extracted_data['subjects'].append({
                        'name': str(subject_col).strip(),
                        'code': ExcelDataExtractor._generate_subject_code(str(subject_col))
                    })
            
            # Extraer estudiantes y calificaciones
            for index, row in df.iterrows():
                # Datos del estudiante
                student_id = str(row[id_col]).strip()
                last_name = str(row[lastname_col]).strip() if pd.notna(row[lastname_col]) else ''
                first_name = str(row[firstname_col]).strip() if pd.notna(row[firstname_col]) else ''
                
                if not student_id or student_id.lower() == 'nan':
                    continue
                
                student_data = {
                    'student_id': student_id,
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': ExcelDataExtractor._generate_email(first_name, last_name, student_id),
                    'grades': {}
                }
                
                # Extraer calificaciones
                for i, subject_col in enumerate(subject_cols):
                    if pd.notna(subject_col):
                        grade_value = row[subject_col]
                        if pd.notna(grade_value):
                            try:
                                # Intentar convertir a número
                                numeric_grade = float(grade_value)
                                if 0 <= numeric_grade <= 100:  # Validar rango
                                    student_data['grades'][str(subject_col).strip()] = numeric_grade
                            except (ValueError, TypeError):
                                # Si no es numérico, guardar como texto
                                student_data['grades'][str(subject_col).strip()] = str(grade_value).strip()
                
                extracted_data['students'].append(student_data)
            
            return extracted_data
            
        except Exception as e:
            raise Exception(f"Error al procesar archivo Excel: {str(e)}")
    
    @staticmethod
    def _generate_subject_code(subject_name):
        """Genera un código para la materia basado en su nombre"""
        # Tomar las primeras letras de cada palabra
        words = subject_name.upper().split()
        if len(words) == 1:
            return words[0][:4]
        else:
            return ''.join(word[0] for word in words if word)[:4]
    
    @staticmethod
    def _generate_email(first_name, last_name, student_id):
        """Genera un email básico para el estudiante"""
        # Limpiar nombres
        first_clean = re.sub(r'[^a-zA-Z]', '', first_name.lower())
        last_clean = re.sub(r'[^a-zA-Z]', '', last_name.lower())
        
        if first_clean and last_clean:
            return f"{first_clean}.{last_clean}@estudiante.edu"
        elif first_clean:
            return f"{first_clean}.{student_id}@estudiante.edu"
        else:
            return f"estudiante.{student_id}@estudiante.edu"
    
    @staticmethod
    def analyze_excel_structure(file_path):
        """Analiza la estructura de un archivo Excel y retorna información sobre las hojas y columnas"""
        
        try:
            # Cargar archivo para obtener nombres de hojas
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            
            analysis = {
                'sheets': [],
                'recommended_sheet': None
            }
            
            for sheet_name in sheet_names:
                # Leer primeras filas de cada hoja
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
                
                sheet_info = {
                    'name': sheet_name,
                    'columns': df.columns.tolist(),
                    'row_count': len(df),
                    'has_student_data': ExcelDataExtractor._detect_student_data(df.columns.tolist())
                }
                
                analysis['sheets'].append(sheet_info)
                
                # Recomendar hoja que parece tener datos de estudiantes
                if sheet_info['has_student_data'] and not analysis['recommended_sheet']:
                    analysis['recommended_sheet'] = sheet_name
            
            return analysis
            
        except Exception as e:
            raise Exception(f"Error al analizar estructura del archivo: {str(e)}")
    
    @staticmethod
    def _detect_student_data(columns):
        """Detecta si las columnas parecen contener datos de estudiantes"""
        
        student_indicators = [
            'id', 'cedula', 'estudiante', 'alumno', 'nombre', 'apellido',
            'first', 'last', 'student', 'name', 'surname'
        ]
        
        column_text = ' '.join(str(col).lower() for col in columns)
        
        matches = sum(1 for indicator in student_indicators if indicator in column_text)
        
        return matches >= 2  # Al menos 2 indicadores
    
    @staticmethod
    def validate_excel_format(file_path, expected_columns=None):
        """Valida que el archivo Excel tenga el formato esperado"""
        
        try:
            df = pd.read_excel(file_path, nrows=1)
            
            validation_result = {
                'is_valid': True,
                'errors': [],
                'warnings': [],
                'column_count': len(df.columns),
                'detected_columns': df.columns.tolist()
            }
            
            # Validaciones básicas
            if len(df.columns) < 3:
                validation_result['is_valid'] = False
                validation_result['errors'].append("El archivo debe tener al menos 3 columnas")
            
            # Validar columnas esperadas si se proporcionan
            if expected_columns:
                missing_columns = set(expected_columns) - set(df.columns)
                if missing_columns:
                    validation_result['warnings'].append(f"Columnas faltantes: {', '.join(missing_columns)}")
            
            # Verificar si hay datos
            full_df = pd.read_excel(file_path)
            if len(full_df) == 0:
                validation_result['is_valid'] = False
                validation_result['errors'].append("El archivo no contiene datos")
            
            validation_result['row_count'] = len(full_df)
            
            return validation_result
            
        except Exception as e:
            return {
                'is_valid': False,
                'errors': [f"Error al validar archivo: {str(e)}"],
                'warnings': [],
                'column_count': 0,
                'detected_columns': [],
                'row_count': 0
            }

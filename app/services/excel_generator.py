from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from app.models.templates import ExcelTemplate, TemplateCell, TemplateStyle
import json
import os

class ExcelGenerator:
    """Generador base de archivos Excel usando plantillas"""
    
    @staticmethod
    def generate_from_template(template_id, data):
        """Genera un Excel basado en una plantilla y datos proporcionados"""
        
        template = ExcelTemplate.query.get_or_404(template_id)
        
        if template.file_path and os.path.exists(template.file_path):
            # Cargar plantilla existente
            wb = load_workbook(template.file_path)
            ws = wb.active
        else:
            # Crear nuevo workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte"
        
        # Aplicar datos din√°micos
        ExcelGenerator._apply_dynamic_data(ws, template_id, data)
        
        # Aplicar estilos
        ExcelGenerator._apply_template_styles(ws, template_id)
        
        return wb
    
    @staticmethod
    def _apply_dynamic_data(ws, template_id, data):
        """Aplica datos din√°micos a las celdas marcadas como din√°micas"""
        
        dynamic_cells = TemplateCell.query.filter_by(
            template_id=template_id,
            cell_type='dynamic'
        ).all()
        
        for cell in dynamic_cells:
            value = ExcelGenerator._get_data_value(data, cell.content_type)
            if value is not None:
                ws[cell.cell_address] = value
    
    @staticmethod
    def _get_data_value(data, content_type):
        """Obtiene el valor de los datos basado en el tipo de contenido"""
        
        if not content_type:
            return None
        
        # Mapeo de tipos de contenido a datos
        content_mappings = {
            'section_name': lambda d: d.get('section', {}).get('name', ''),
            'grade_name': lambda d: d.get('section', {}).get('grade', {}).get('name', ''),
            'period_name': lambda d: d.get('period', {}).get('name', ''),
            'student_name': lambda d: d.get('student', {}).get('full_name', ''),
            'student_id': lambda d: d.get('student', {}).get('id', ''),
            'current_date': lambda d: d.get('current_date', ''),
            'average': lambda d: d.get('average', ''),
        }
        
        if content_type in content_mappings:
            return content_mappings[content_type](data)
        
        return None
    
    @staticmethod
    def _apply_template_styles(ws, template_id):
        """Aplica estilos de la plantilla"""
        
        # Aplicar estilos de celdas
        cells = TemplateCell.query.filter_by(template_id=template_id).all()
        
        for cell in cells:
            if cell.style_config:
                try:
                    style_config = json.loads(cell.style_config)
                    ExcelGenerator._apply_cell_style(ws[cell.cell_address], style_config)
                except json.JSONDecodeError:
                    continue
        
        # Aplicar estilos de rangos
        styles = TemplateStyle.query.filter_by(template_id=template_id).all()
        
        for style in styles:
            ExcelGenerator._apply_range_style(ws, style)
    
    @staticmethod
    def _apply_cell_style(cell, style_config):
        """Aplica estilo a una celda espec√≠fica"""
        
        print(f"üé® DEBUG ExcelGenerator: Aplicando estilo a {cell.coordinate}")
        print(f"   üìÑ Config recibida: {style_config}")
        
        # Fuente
        if 'font' in style_config:
            font_config = style_config['font']
            font_color = font_config.get('color', '000000')
            print(f"   üìù Font color: {font_color}")
            
            cell.font = Font(
                name=font_config.get('name', 'Arial'),
                size=font_config.get('size', 11),
                bold=font_config.get('bold', False),
                italic=font_config.get('italic', False),
                color=font_color
            )
        
        # Relleno
        if 'fill' in style_config:
            fill_config = style_config['fill']
            fill_color = fill_config.get('color', 'FFFFFF')
            print(f"   üé® Fill color: {fill_color}")
            
            # SOLO aplicar si no es blanco
            if fill_color.upper() != 'FFFFFF':
                cell.fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type='solid'
                )
                print(f"   ‚úÖ Fill aplicado")
            else:
                print(f"   ‚ö™ Fill blanco ignorado")
        
        # Alineaci√≥n
        if 'alignment' in style_config:
            align_config = style_config['alignment']
            cell.alignment = Alignment(
                horizontal=align_config.get('horizontal', 'left'),
                vertical=align_config.get('vertical', 'top'),
                wrap_text=align_config.get('wrap_text', False)
            )
    
    @staticmethod
    def _apply_range_style(ws, style):
        """Aplica estilos a rangos de celdas"""
        
        # Ancho de columna
        if style.column_width and ':' in style.range_address:
            col_range = style.range_address.split(':')
            if col_range[0] == col_range[1]:  # Columna √∫nica
                ws.column_dimensions[col_range[0]].width = style.column_width
        
        # Alto de fila
        if style.row_height and style.range_address.isdigit():
            ws.row_dimensions[int(style.range_address)].height = style.row_height

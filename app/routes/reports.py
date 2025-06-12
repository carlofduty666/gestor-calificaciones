from flask import Blueprint, render_template, request, send_file, flash, jsonify, redirect, url_for
from flask_login import login_required, current_user
from app.models.users import User
from app.models.academic import AcademicYear, Period, Grade, Section, Subject, Student, Teacher, TeacherAssignment
from app.models.grades import GradeType, StudentGrade, FinalGrade
from app import db
import pandas as pd
from io import BytesIO
import tempfile
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from app.services.multi_sheet_excel_generator import MultiSheetExcelGenerator
import os
import re
from datetime import datetime

reports = Blueprint('reports', __name__)

@reports.before_request
def check_access():
    if not current_user.is_authenticated:
        return redirect(url_for('auth.login'))

@reports.route('/')
@login_required
def index():
    # Obtener año académico activo
    active_year = AcademicYear.query.filter_by(is_active=True).first()
    
    if not active_year:
        flash('No hay un año académico activo', 'warning')
        return render_template('reports/index.html', 
                             active_year=None, 
                             periods=[], 
                             sections=[],
                             academic_years=[],
                             grades=[])
    
    # Obtener períodos del año activo
    periods = active_year.periods.order_by(Period.start_date).all()
    
    # Obtener todas las secciones
    sections = Section.query.join(Grade).order_by(Grade.name, Section.name).all()
    
    # Obtener todos los años académicos para el filtro
    academic_years = AcademicYear.query.order_by(AcademicYear.start_date.desc()).all()
    
    # Obtener todos los grados
    grades = Grade.query.order_by(Grade.name).all()
    
    return render_template('reports/index.html',
                         active_year=active_year,
                         periods=periods,
                         sections=sections,
                         academic_years=academic_years,
                         grades=grades)

@reports.route('/section/<int:section_id>/period/<int:period_id>')
@login_required
def section_report(section_id, period_id):
    section = Section.query.get_or_404(section_id)
    period = Period.query.get_or_404(period_id)
    
    # Verificar acceso
    if not current_user.is_admin():
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        
        if not teacher:
            flash('No se encontró un perfil de profesor para este usuario', 'warning')
            return redirect(url_for('auth.logout'))
        
        # Verificar si el profesor tiene asignaciones en esta sección
        assignment = TeacherAssignment.query.filter_by(
            teacher_id=teacher.id,
            section_id=section_id,
            academic_year_id=period.academic_year_id
        ).first()
        
        if not assignment:
            flash('No tienes permiso para ver esta sección', 'danger')
            return redirect(url_for('reports.index'))
    
    # Obtener estudiantes de la sección
    students = Student.query.filter_by(
        section_id=section_id,
        is_active=True
    ).order_by(Student.last_name).all()
    
    # Obtener asignaturas para esta sección (a través de las asignaciones)
    subject_ids = db.session.query(TeacherAssignment.subject_id).filter_by(
        section_id=section_id,
        academic_year_id=period.academic_year_id
    ).distinct().all()
    
    subject_ids = [s[0] for s in subject_ids]
    subjects = Subject.query.filter(Subject.id.in_(subject_ids)).order_by(Subject.name).all()
    
    # Obtener calificaciones finales
    grades_data = {}
    for student in students:
        grades_data[student.id] = {}
        for subject in subjects:
            final_grade = FinalGrade.query.filter_by(
                student_id=student.id,
                subject_id=subject.id,
                period_id=period.id
            ).first()
            
            if final_grade:
                grades_data[student.id][subject.id] = final_grade.value
    
    return render_template('reports/section_report.html',
                          title=f'Reporte - {section.grade.name}{section.name} - {period.name}',
                          section=section,
                          period=period,
                          students=students,
                          subjects=subjects,
                          grades_data=grades_data)

@reports.route('/student/<int:student_id>/period/<int:period_id>')
@login_required
def student_report(student_id, period_id):
    student = Student.query.get_or_404(student_id)
    period = Period.query.get_or_404(period_id)
    
    # Verificar acceso
    if not current_user.is_admin():
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        
        if not teacher:
            flash('No se encontró un perfil de profesor para este usuario', 'warning')
            return redirect(url_for('auth.logout'))
        
        # Verificar si el profesor tiene asignaciones en la sección del estudiante
        assignment = TeacherAssignment.query.filter_by(
            teacher_id=teacher.id,
            section_id=student.section_id,
            academic_year_id=period.academic_year_id
        ).first()
        
        if not assignment:
            flash('No tienes permiso para ver este estudiante', 'danger')
            return redirect(url_for('reports.index'))
    
    # Obtener asignaturas para la sección del estudiante
    subject_ids = db.session.query(TeacherAssignment.subject_id).filter_by(
        section_id=student.section_id,
        academic_year_id=period.academic_year_id
    ).distinct().all()
    
    subject_ids = [s[0] for s in subject_ids]
    subjects = Subject.query.filter(Subject.id.in_(subject_ids)).order_by(Subject.name).all()
    
    # Obtener calificaciones detalladas
    grades_data = {}
    final_grades = {}
    
    for subject in subjects:
        # Obtener tipos de calificación para esta asignatura
        grade_types = GradeType.query.filter_by(
            subject_id=subject.id,
            period_id=period.id
        ).order_by(GradeType.name).all()
        
        grades_data[subject.id] = {}
        
        for grade_type in grade_types:
            grade = StudentGrade.query.filter_by(
                student_id=student.id,
                subject_id=subject.id,
                grade_type_id=grade_type.id,
                period_id=period.id
            ).first()
            
            if grade:
                grades_data[subject.id][grade_type.id] = {
                    'name': grade_type.name,
                    'value': grade.value,
                    'weight': grade_type.weight
                }
        
        # Obtener calificación final
        final_grade = FinalGrade.query.filter_by(
            student_id=student.id,
            subject_id=subject.id,
            period_id=period.id
        ).first()
        
        if final_grade:
            final_grades[subject.id] = final_grade.value
    
    return render_template('reports/student_report.html',
                          title=f'Reporte - {student.first_name} {student.last_name} - {period.name}',
                          student=student,
                          period=period,
                          subjects=subjects,
                          grades_data=grades_data,
                          final_grades=final_grades)

@reports.route('/export/excel/section/<int:section_id>/period/<int:period_id>')
@login_required
def export_section_excel(section_id, period_id):
    section = Section.query.get_or_404(section_id)
    period = Period.query.get_or_404(period_id)
    
    # Verificar acceso (mismo código que en section_report)
    if not current_user.is_admin():
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        
        if not teacher:
            flash('No se encontró un perfil de profesor para este usuario', 'warning')
            return redirect(url_for('auth.logout'))
        
        assignment = TeacherAssignment.query.filter_by(
            teacher_id=teacher.id,
            section_id=section_id,
            academic_year_id=period.academic_year_id
        ).first()
        
        if not assignment:
            flash('No tienes permiso para exportar esta sección', 'danger')
            return redirect(url_for('reports.index'))
    
    # Obtener estudiantes de la sección
    students = Student.query.filter_by(
        section_id=section_id,
        is_active=True
    ).order_by(Student.last_name).all()
    
    # Obtener asignaturas para esta sección
    subject_ids = db.session.query(TeacherAssignment.subject_id).filter_by(
        section_id=section_id,
        academic_year_id=period.academic_year_id
    ).distinct().all()
    
    subject_ids = [s[0] for s in subject_ids]
    subjects = Subject.query.filter(Subject.id.in_(subject_ids)).order_by(Subject.name).all()
    
    # Crear DataFrame para Excel
    data = []
    for student in students:
        row = {
            'ID': student.student_id,
            'Apellidos': student.last_name,
            'Nombres': student.first_name
        }
        
        # Añadir calificaciones por asignatura
        for subject in subjects:
            final_grade = FinalGrade.query.filter_by(
                student_id=student.id,
                subject_id=subject.id,
                period_id=period.id
            ).first()
            
            if final_grade:
                row[subject.name] = final_grade.value
            else:
                row[subject.name] = ''
        
        data.append(row)
    
    # Crear Excel
    df = pd.DataFrame(data)
    
    # Crear archivo Excel en memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=f'{section.grade.name}{section.name}', index=False)
    
    output.seek(0)
    
    # Generar nombre de archivo
    filename = f'Calificaciones_{section.grade.name}{section.name}_{period.name}.xlsx'
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@reports.route('/export/pdf/section/<int:section_id>/period/<int:period_id>')
@login_required
def export_section_pdf(section_id, period_id):
    section = Section.query.get_or_404(section_id)
    period = Period.query.get_or_404(period_id)
    
    # Verificar acceso (mismo código que en section_report)
    if not current_user.is_admin():
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        
        if not teacher:
            flash('No se encontró un perfil de profesor para este usuario', 'warning')
            return redirect(url_for('auth.logout'))
        
        assignment = TeacherAssignment.query.filter_by(
            teacher_id=teacher.id,
            section_id=section_id,
            academic_year_id=period.academic_year_id
        ).first()
        
        if not assignment:
            flash('No tienes permiso para exportar esta sección', 'danger')
            return redirect(url_for('reports.index'))
    
    # Obtener estudiantes de la sección
    students = Student.query.filter_by(
        section_id=section_id,
        is_active=True
    ).order_by(Student.last_name).all()
    
    # Obtener asignaturas para esta sección
    subject_ids = db.session.query(TeacherAssignment.subject_id).filter_by(
        section_id=section_id,
        academic_year_id=period.academic_year_id
    ).distinct().all()
    
    subject_ids = [s[0] for s in subject_ids]
    subjects = Subject.query.filter(Subject.id.in_(subject_ids)).order_by(Subject.name).all()
    
    # Crear archivo temporal
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
        temp_filename = temp_file.name
    
    # Crear PDF con ReportLab
    doc = SimpleDocTemplate(temp_filename, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    subtitle_style = styles['Heading2']
    
    # Título
    elements.append(Paragraph(f"Reporte de Calificaciones", title_style))
    elements.append(Paragraph(f"Grado: {section.grade.name} Sección: {section.name}", subtitle_style))
    elements.append(Paragraph(f"Período: {period.name}", subtitle_style))
    
    # Datos para la tabla
    data = [['ID', 'Apellidos', 'Nombres'] + [subject.name for subject in subjects]]
    
    for student in students:
        row = [
            student.student_id,
            student.last_name,
            student.first_name
        ]
        
        # Añadir calificaciones por asignatura
        for subject in subjects:
            final_grade = FinalGrade.query.filter_by(
                student_id=student.id,
                subject_id=subject.id,
                period_id=period.id
            ).first()
            
            if final_grade:
                row.append(str(final_grade.value))
            else:
                row.append('')
        
        data.append(row)
    
    # Crear tabla
    table = Table(data)
    
    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    
    table.setStyle(style)
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    # Generar nombre de archivo
    filename = f'Calificaciones_{section.grade.name}{section.name}_{period.name}.pdf'
    
    return send_file(
        temp_filename,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )

@reports.route('/export/pdf/student/<int:student_id>/period/<int:period_id>')
@login_required
def export_student_pdf(student_id, period_id):
    student = Student.query.get_or_404(student_id)
    period = Period.query.get_or_404(period_id)
    
    # Verificar acceso (mismo código que en student_report)
    if not current_user.is_admin():
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        
        if not teacher:
            flash('No se encontró un perfil de profesor para este usuario', 'warning')
            return redirect(url_for('auth.logout'))
        
        assignment = TeacherAssignment.query.filter_by(
            teacher_id=teacher.id,
            section_id=student.section_id,
            academic_year_id=period.academic_year_id
        ).first()
        
        if not assignment:
            flash('No tienes permiso para exportar este estudiante', 'danger')
            return redirect(url_for('reports.index'))
    
    # Obtener asignaturas para la sección del estudiante
    subject_ids = db.session.query(TeacherAssignment.subject_id).filter_by(
        section_id=student.section_id,
        academic_year_id=period.academic_year_id
    ).distinct().all()
    
    subject_ids = [s[0] for s in subject_ids]
    subjects = Subject.query.filter(Subject.id.in_(subject_ids)).order_by(Subject.name).all()
    
    # Crear archivo temporal
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
        temp_filename = temp_file.name
    
    # Crear PDF con ReportLab
    doc = SimpleDocTemplate(temp_filename, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    subtitle_style = styles['Heading2']
    normal_style = styles['Normal']
    
    # Título
    elements.append(Paragraph(f"Boleta de Calificaciones", title_style))
    elements.append(Paragraph(f"Estudiante: {student.first_name} {student.last_name}", subtitle_style))
    elements.append(Paragraph(f"C.I: {student.student_id}", normal_style))
    elements.append(Paragraph(f"Grado: {student.section.grade.name} Sección: {student.section.name}", normal_style))
    elements.append(Paragraph(f"Período: {period.name}", normal_style))
    elements.append(Paragraph(" ", normal_style))  # Espacio
    
    # Datos para la tabla
    data = [['Asignatura', 'Calificación']]
    
    for subject in subjects:
        final_grade = FinalGrade.query.filter_by(
            student_id=student.id,
            subject_id=subject.id,
            period_id=period.id
        ).first()
        
        if final_grade:
            data.append([subject.name, str(final_grade.value)])
        else:
            data.append([subject.name, ''])
    
    # Calcular promedio
    final_grades = [row[1] for row in data[1:] if row[1]]
    if final_grades:
        try:
            average = sum(float(grade) for grade in final_grades) / len(final_grades)
            data.append(['Promedio', f'{average:.2f}'])
        except ValueError:
            data.append(['Promedio', 'N/A'])
    
    # Crear tabla
    table = Table(data)
    
    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    
    table.setStyle(style)
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    # Generar nombre de archivo
    filename = f'Boleta_{student.last_name}_{student.first_name}_{period.name}.pdf'
    
    return send_file(
        temp_filename,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )

# Agregar esta nueva ruta
@reports.route('/export/multi-sheet/section/<int:section_id>/period/<int:period_id>')
@login_required
def export_multi_sheet_section(section_id, period_id):
    """Exporta un Excel con múltiples hojas, una por cada materia de la sección"""
    
    section = Section.query.get_or_404(section_id)
    period = Period.query.get_or_404(period_id)
    
    # Verificar acceso
    if not current_user.is_admin():
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        if not teacher:
            flash('No se encontró un perfil de profesor para este usuario', 'warning')
            return redirect(url_for('auth.logout'))
        
        assignment = TeacherAssignment.query.filter_by(
            teacher_id=teacher.id,
            section_id=section_id,
            academic_year_id=period.academic_year_id
        ).first()
        
        if not assignment:
            flash('No tienes permiso para exportar esta sección', 'danger')
            return redirect(url_for('reports.index'))
    
    template_id = request.args.get('template_id')
    
    try:
        # Generar Excel con múltiples hojas
        wb = MultiSheetExcelGenerator.generate_section_multi_subject_report(
            section_id, period.id, template_id
        )
        
        # Crear archivo en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'Reporte_Completo_{section.grade.name}{section.name}_{period.name}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error al generar reporte: {str(e)}', 'danger')
        return redirect(url_for('reports.section_report', section_id=section_id, period_id=period_id))
    
# Agregar esta nueva ruta al final del archivo
@reports.route('/export/template/<int:template_id>/section/<int:section_id>/period/<int:period_id>')
@login_required
def export_with_template(template_id, section_id, period_id):
    """Exporta usando una plantilla con estilos preservados"""
    
    from app.models.templates import ExcelTemplate, TemplateCell
    from app.services.template_mapper import TemplateMapper
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    import json
    
    # Verificar acceso (mismo código que otras funciones)
    section = Section.query.get_or_404(section_id)
    period = Period.query.get_or_404(period_id)
    template = ExcelTemplate.query.get_or_404(template_id)
    
    if not current_user.is_admin():
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        if not teacher:
            flash('No se encontró un perfil de profesor para este usuario', 'warning')
            return redirect(url_for('auth.logout'))
        
        assignment = TeacherAssignment.query.filter_by(
            teacher_id=teacher.id,
            section_id=section_id,
            academic_year_id=period.academic_year_id
        ).first()
        
        if not assignment:
            flash('No tienes permiso para exportar esta sección', 'danger')
            return redirect(url_for('reports.index'))
    
    # Cargar plantilla
    if not template.file_path or not os.path.exists(template.file_path):
        flash('Archivo de plantilla no encontrado', 'danger')
        return redirect(url_for('reports.index'))
    
    wb = load_workbook(template.file_path)
    ws = wb.active
    
    # Obtener estudiantes
    students = Student.query.filter_by(
        section_id=section_id,
        is_active=True
    ).order_by(Student.last_name).all()
    
    # Obtener configuración de celdas
    template_cells = TemplateCell.query.filter_by(template_id=template_id).all()
    
    # Preparar contexto
    context = {
        'section': section,
        'grade': section.grade,
        'period': period,
        'period_id': period_id,
        'current_date': datetime.now().strftime('%d/%m/%Y')
    }
    
    # Encontrar la fila donde empiezan los datos de estudiantes
    data_start_row = None
    for cell in template_cells:
        if cell.cell_type == 'data':
            row_num = int(re.search(r'\d+', cell.cell_address).group())
            if data_start_row is None or row_num < data_start_row:
                data_start_row = row_num
    
    if data_start_row is None:
        data_start_row = 2  # Por defecto empezar en fila 2
    
    # Llenar datos de estudiantes
    for student_index, student in enumerate(students):
        current_row = data_start_row + student_index
        
        # Llenar cada celda de datos para este estudiante
        for cell in template_cells:
            if cell.cell_type == 'data':
                # Calcular nueva posición
                col_letter = re.search(r'[A-Z]+', cell.cell_address).group()
                new_address = f"{col_letter}{current_row}"
                
                # Obtener valor
                value = TemplateMapper.get_value_for_cell(cell.data_type, student, context)
                
                if value is not None:
                    ws[new_address] = value
                    
                    # Aplicar estilo original si existe
                    if cell.style_config:
                        try:
                            style_config = json.loads(cell.style_config)
                            _apply_cell_style(ws[new_address], style_config)
                        except:
                            pass
    
    # Crear archivo en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'Reporte_{template.name}_{section.grade.name}{section.name}_{period.name}.xlsx'
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def _apply_cell_style(cell, style_config):
    """Aplica estilo a una celda"""
    from openpyxl.styles import Font, PatternFill, Alignment
    
    if 'font' in style_config:
        font_config = style_config['font']
        cell.font = Font(
            name=font_config.get('name', 'Arial'),
            size=font_config.get('size', 11),
            bold=font_config.get('bold', False),
            italic=font_config.get('italic', False)
        )
    
    if 'fill' in style_config:
        fill_config = style_config['fill']
        cell.fill = PatternFill(
            start_color=fill_config.get('color', 'FFFFFF'),
            end_color=fill_config.get('color', 'FFFFFF'),
            fill_type='solid'
        )
    
    if 'alignment' in style_config:
        align_config = style_config['alignment']
        cell.alignment = Alignment(
            horizontal=align_config.get('horizontal', 'left'),
            vertical=align_config.get('vertical', 'top')
        )

# Agregar esta ruta también al final del archivo

@reports.route('/templates')
@login_required
def templates_list():
    """Lista las plantillas disponibles para generar reportes"""
    
    from app.models.templates import ExcelTemplate
    
    # Obtener año académico activo
    active_year = AcademicYear.query.filter_by(is_active=True).first()
    
    if not active_year:
        flash('No hay un año académico activo', 'warning')
        return render_template('reports/templates.html', 
                             active_year=None, 
                             periods=[], 
                             sections=[],
                             templates=[])
    
    # Obtener datos necesarios
    periods = active_year.periods.order_by(Period.start_date).all()
    sections = Section.query.join(Grade).order_by(Grade.name, Section.name).all()
    templates = ExcelTemplate.query.filter_by(is_active=True).all()
    
    return render_template('reports/templates.html',
                         active_year=active_year,
                         periods=periods,
                         sections=sections,
                         templates=templates)

@reports.route('/template/preview/<int:template_id>')
@login_required
def template_preview(template_id):
    """Vista previa de una plantilla"""
    
    from app.services.template_service import TemplateService
    
    try:
        preview_data = TemplateService.generate_preview(template_id)
        return jsonify({
            'success': True,
            'preview': preview_data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })
